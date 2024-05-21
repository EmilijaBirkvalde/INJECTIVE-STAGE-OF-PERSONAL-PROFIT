import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import time



def iegut_un_saglabat_cenu():
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    url = "https://www.diadata.org/app/price/asset/Ethereum/0xe28b3B32B6c345A34Ff64674606124Dd5Aceca30/"
    driver.get(url)
    time.sleep(5)

    find = driver.find_element(By.CLASS_NAME, "btn.green")
    find.click()
    time.sleep(1)

    cena_elements = driver.find_elements(By.CLASS_NAME, "price.text-14-28.strong")
    pašreizējā_cena = cena_elements[0].text if cena_elements else 'N/A'

    driver.quit()
    return pašreizējā_cena




def saglabat_datus_excel(dati):
    darba_grāmata = load_workbook('crypto_dati.xlsx')

    darblapa = darba_grāmata.active
    tagad = datetime.now()
    datums_laiks = tagad.strftime('%Y-%m-%d %H:%M:%S')

    darblapa.append([datums_laiks, dati['cena'], dati['vienu_skaits'], dati['aprēķins']])

    darba_grāmata.save('crypto_dati.xlsx')




def aprēķinat_profitu_zaudējumu(pašreizējā_cena, piederīgie_vienu_skaits, iepriekšējā_cena):
    try:
        iepriekšējā_cena = float(iepriekšējā_cena.replace('$', '').replace(',', ''))
        pašreizējā_cena = float(pašreizējā_cena.replace('$', '').replace(',', ''))
        profit_loss = (pašreizējā_cena - iepriekšējā_cena) * piederīgie_vienu_skaits
        return round(profit_loss, 2)
    except ValueError:
        return 'N/A'




pašreizējā_cena = iegut_un_saglabat_cenu()
iepriekšējā_cena = ''
try:
    darba_grāmata = load_workbook('crypto_dati.xlsx')
    darblapa = darba_grāmata.active
    iepriekšējā_cena = darblapa.cell(row=darblapa.max_row, column=2).value
except FileNotFoundError:
    print('Iepriekšējā cenu nevar nolasīt, pirmo reizi izpildot programmu.')


vienu_skaits = float(input('Ievadiet injekciju vienību skaitu, kas jums pieder: '))

profit_loss = aprēķinat_profitu_zaudējumu(pašreizējā_cena, vienu_skaits, iepriekšējā_cena)

saglabat_datus = {
    'cena': pašreizējā_cena,
    'vienu_skaits': vienu_skaits,
    'aprēķins': profit_loss
}

saglabat_datus_excel(saglabat_datus)

print(f'Dati saglabāti Excel failā ("crypto_dati.xlsx").')
